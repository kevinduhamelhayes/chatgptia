import openpyxl

wb = openpyxl.load_workbook('example.xlsx')

print(wb.sheetnames)
print(wb["Hoja1"])
wb.create_sheet("Hoja2")
print(wb.sheetnames)
hoja2 = wb["Hoja2"]
hoja2.title = "HojadePrueba"

celda = hoja1["A1"]
print(celda.value)
celda.value = "Hola"
celda2 = hoja1.cell(row=1, column=2)
print(celda2.value)
print(hoja1.max_row)
for fila in hoja1:
    for celda in fila:
        print(celda.value)
columna = hoja1["A"]
fila = hoja1[1]
hoja.append([1,2,3])
print(hoja1.max_row)

hoja.delete_rows(1)

